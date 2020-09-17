import tkinter as tk
import tkinter.ttk as ttk
from openpyxl import *
from tkinter.messagebox import showinfo

win = tk.Tk()
win.title("Users Registration Form")

def save():
	firstName = entry.get()
	lastName = entry1.get()
	age = entry2.get()
	wb = Workbook()
	ws = wb.active
	ws['A1'] = "First Name"
	ws['B1'] = "Last Name"
	ws['C1'] = "Age"
	ws['A2'] = firstName
	ws['B2'] = lastName
	ws['C2'] = age
	wb.save(r'C:\Users\zinou\Documents\users.xlsx')
	showinfo("Saved", "user data saved successfully")

def clear():
	entry.delete(0,tk.END)
	entry1.delete(0,tk.END)
	entry2.delete(0,tk.END)

label = tk.Label(win,text="First Name: ")
label.grid(row=0,column=0,padx=8,pady=8)
entry = tk.Entry(win)
entry.grid(row=0,column=1,padx=8,pady=8)

label1 = tk.Label(win,text="Last Name: ")
label1.grid(row=1,column=0,padx=8,pady=8)
entry1 = tk.Entry(win)
entry1.grid(row=1,column=1,padx=8, pady=8)

label2 = tk.Label(win,text="Age: ")
label2.grid(row=2,column=0,padx=8,pady=8)
entry2 = tk.Entry(win)
entry2.grid(row=2,column=1,padx=8,pady=8)

buttonReg = ttk.Button(win,text="Register",command=save)
buttonReg.grid(row=3,column=0,padx=8,pady=8)

buttonClear = ttk.Button(win,text="clear",command=clear)
buttonClear.grid(row=3,column=1,padx=8,pady=8)

win.mainloop()
